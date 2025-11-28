"""Utilities for exporting the generated schedule to Excel."""

from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from .scheduler import ScheduledItem


def export_schedule_to_excel(rows: Iterable[ScheduledItem], path: Path) -> None:
    """Write schedule rows to an Excel workbook.

    Args:
        rows: Iterable of :class:`ScheduledItem` values to export.
        path: Location for the resulting ``.xlsx`` file. Parent directories are
            created automatically.
    """

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Schedule"

    headers = ["Level", "Week", "Day", "Activity", "Module", "Duration(min)", "Goal"]
    sheet.append(headers)

    for item in rows:
        sheet.append(
            [
                item.level,
                item.week,
                item.day,
                item.activity,
                item.module,
                item.duration_minutes,
                item.goal,
            ]
        )

    for column_index, header in enumerate(headers, start=1):
        max_length = len(header)
        for cell in sheet.iter_cols(
            min_col=column_index,
            max_col=column_index,
            min_row=2,
            max_row=sheet.max_row,
        ):
            for value_cell in cell:
                if value_cell.value is not None:
                    max_length = max(max_length, len(str(value_cell.value)))
        sheet.column_dimensions[get_column_letter(column_index)].width = min(max_length + 2, 60)

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


__all__ = ["export_schedule_to_excel"]
